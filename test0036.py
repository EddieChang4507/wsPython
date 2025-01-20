# 讀取 EXCEL 內容
# 
# pip install openpyxl
# 

import openpyxl

file_path = 'C:\\test'
file_xlsx_A = '整合測試劇本(SoftPOS)第二階段需求.xlsx'


wb = openpyxl.load_workbook(f'{file_path}\\{file_xlsx_A}')     # 開啟 Excel 檔案

names = wb.sheetnames    # 讀取 Excel 裡所有工作表名稱
print(names)

for item in names:
    s1 = wb[item]            # 取得工作表名稱為「工作表1」的內容
    # s2 = wb.active           # 取得開啟試算表後立刻顯示的工作表 ( 範例為工作表 2 )

    # 印出 title ( 工作表名稱 )、max_row 最大列數、max_column 最大行數
    print(s1.title, s1.max_row, s1.max_column)
    # print(s2.title, s2.max_row, s2.max_column)
    
    # print(s1['A1'].value)        # 取出 A1 的內容
    # print(s1.cell(1, 1).value)   # 等同取出 A1 的內容
    
    def get_values(sheet):
        arr = []                      # 第一層串列
        for row in sheet:
            arr2 = []                 # 第二層串列
            for column in row:
                arr2.append(column.value)  # 寫入內容
            arr.append(arr2)
        return arr

    print(get_values(s1))       # 印出工作表 1 所有內容
